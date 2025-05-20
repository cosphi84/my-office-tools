'''
csopp.py
Tools untuk melakukan ekstraksi pencapaian pekerjan, berdasarkan target KPI CS
Sumber data:
1. Ots.XLS : File csv job orders status Outstanding
2. Completed.XLS : File csv completed data
'''
import pandas as pd
import numpy as np
import os
from pathlib import Path
import warnings
from datetime import datetime
from functools import reduce
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.colors import WHITE
from file_io import check_file, remove_file, load_file

# Disable deprecated warning
warnings.simplefilter(action="ignore", category=FutureWarning)
warnings.simplefilter(action="ignore", category=UserWarning)


# File path constants
FILE_CONFIG = Path("Config/csopp.xlsx")
FILE_OTS = Path('DataSource/ots.xlsx')
FILE_COMPLETED = Path("DataSource/completed.xlsx")
FILE_OTS_ALT = Path('DataSource/ots.csv')
FILE_COMPLETED_ALT = Path("DataSource/completed.csv")
FILE_RESULT = Path("Result/Result.xlsx")
FILE_DETAIL = Path("Result/Detail.xlsx")
FILE_ERROR = Path("Result/Error.xlsx")
FILE_TEKNISI = Path("Result/JadwalTeknisi.xlsx")
SHEET_NAMES = {
    "Result": "Result",
    "Prod_C": "Produktifitas Cabang",
    "Prod_D": "Produktifitas SDSS",
    "Prod_R": "Produktifitas SSR"
}


def load_source(path: Path):
    cols = ["Typ", "Notifctn", "Notif.date", "Req. start", "Req. End", "Changed on", "Completion", "PG",  "Mn.wk.ctr", "UserStatus", "List name", "Street", "Telephone", "Material", "Serial number", "Description", "Addit. device data"]
    isOts = False
    if path.name == 'ots.csv' or path.name == 'ots.xlsx':
        isOts = True
    sekarang = datetime.now()

    # Load beberapa config data
    cfg_excel = pd.ExcelFile(FILE_CONFIG)
    with cfg_excel as xls:
        pg = pd.read_excel(xls, "pg")
        mwc = pd.read_excel(xls, "mwc").rename(columns={'Kode':'Mn.wk.ctr', "Teknisi": "Work Center"})
        teknisi = pd.read_excel(xls, "teknisi", usecols=["id", "name"]).rename(columns={"id": "idCSMS", "name":"Teknisi"})
        excludeMwc = pd.read_excel(xls, "exclude", usecols=["mwc"])
    
    df = pd.read_csv(path, usecols=cols, dialect="excel-tab", encoding="utf_16", skiprows=3)
    # exclude Type LR
    df = df[~df["Typ"].isin(["Z8", "ZZ"])]
    # Filter by Status
    df = df[~df["UserStatus"].isin([51, 52])]

    # Convert some stuff
    df["Notif.date"] = pd.to_datetime(df["Notif.date"], dayfirst=True, errors="coerce")
    df["Completion"] = pd.to_datetime(df["Completion"], errors="coerce", format="%d.%m.%Y")
    df["Changed on"] = pd.to_datetime(df["Changed on"], dayfirst=True, errors="coerce")
    df["UserStatus"] = pd.to_numeric(df["UserStatus"], errors="coerce")
    df["PG"] = pd.to_numeric(df["PG"], errors="coerce")
    df["Mn.wk.ctr"] = df["Mn.wk.ctr"].astype(str)
    df["bulan_complet"] = df["Completion"].fillna(sekarang).values.astype("datetime64[M]")
    df["RcvdThisMo"] = (df["Notif.date"] >= df["bulan_complet"]).astype(int)
    df.drop(columns=["bulan_complet"], inplace=True)
    df["Req. End"] = pd.to_datetime(df["Req. End"], errors="coerce", format="%d.%m.%Y")

    # add some stuff    
    if isOts :
        df["isOTS"] = 1
        df["Req. End"] = df["Req. End"].fillna(sekarang)
        df["e_no_req_end"] = 0
    else:
        df["isOTS"] = 0
        df["e_no_req_end"] = df['Req. End'].isna().astype(int)
        df["Req. End"] = df["Req. End"].fillna(sekarang)
        # Buang beberapa Main work yang ingin dibuang
        df = df[~df["Mn.wk.ctr"].isin(excludeMwc)]
    
    
    df["Lo"] = (df["Req. End"] - df["Notif.date"]).dt.days
    df["1D"] = (df["Lo"] <= 1).astype(int)
    df["1W"] = (df["Lo"] <= 7).astype(int)
    df["Cashless"] = df["UserStatus"].isin([94, 95, 96, 98]).astype(int)
    
    # CDR = Cabang, SDSS, SSR
    cdr_kriteria = [
        df["Mn.wk.ctr"].str.startswith("ST"),
        df["Mn.wk.ctr"].str.startswith("SR"),
    ]
    cdr = ["SDSS", "SSR"]
    df["CDR"] = np.select(cdr_kriteria, cdr, default="Cabang")
    df["idCSMS"] = df["Addit. device data"].fillna('').str.split(r"[;:/ ]").str[0]

    pg["PG"] = pd.to_numeric(pg["PG"], errors="coerce")
    mwc["Mn.wk.ctr"] = mwc["Mn.wk.ctr"].astype(str)
    df = df.merge(pg, how="left", on="PG")
    df = df.merge(mwc, how="left", on="Mn.wk.ctr")
    df = df.merge(teknisi, how="left", on="idCSMS")

    return df


def calc_achivement(dfData: pd.DataFrame, GlobalResult : bool = False, Cabang : bool= True):
    idx = ["Regional"] if GlobalResult == True else ["Regional", "Cabang"]
    if GlobalResult :
        idx = ["Regional"]
    elif not GlobalResult and Cabang:
        idx = ["Regional", "Cabang"]
    else:
        idx = ["Regional", "Work Center"]
    pv = pd.pivot_table(dfData, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn": "Total LK"})
    
    # pecah lagi data OTS & data Complete
    dfOTS = dfData[dfData["isOTS"].isin([1])]
    df30 = dfOTS[dfOTS["UserStatus"].isin([30])]
    dfLo = dfOTS[dfOTS["Lo"] >= 60]
    dfComplete = dfData[dfData["isOTS"].isin([0])]
    dfCash = dfData[dfData["UserStatus"].isin([93])]
    # Abaikan type ZX untuk perhitungan Cashless
    dfCash = dfCash[~dfCash["Typ"].isin(["ZX"])]

    # Pivoting
    pv_cmplt = pd.pivot_table(dfComplete, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn": "Komplit"})
    pv_ots = pd.pivot_table(dfOTS, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn": "OTS"})
    pv_1D = pd.pivot_table(dfComplete, index=idx, values=["1D"], aggfunc="sum", fill_value=0, margins=True).rename(columns={"1D": "1 Day"})
    pv_1W = pd.pivot_table(dfComplete, index=idx, values=["1W"], aggfunc="sum", fill_value=0, margins=True).rename(columns={"1W": "1 Week"})
    pv_cash = pd.pivot_table(dfCash, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn":"Cash"})
    pv_cashless = pd.pivot_table(dfComplete, index=idx, values=["Cashless"], aggfunc="sum", fill_value=0, margins=True)
    pv_30 = pd.pivot_table(df30, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn": "STT 30"})
    pv_Lo = pd.pivot_table(dfLo, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn": "LO"})
    pv_TAT = pd.pivot_table(dfComplete, index=idx, values=["Lo"], aggfunc="mean", fill_value=0, margins=True).rename(columns={"Lo": "TAT"})
    pv = reduce(lambda left, right: pd.merge(left, right, on=idx, how="left"), [pv, pv_ots, pv_cmplt, pv_1D, pv_1W, pv_cash, pv_cashless, pv_30, pv_Lo, pv_TAT]).fillna(0)

    pv["Cplt Ratio"] = (pv["Komplit"] / pv["Total LK"])
    pv["LO Ratio"] = (pv["LO"] / pv["OTS"])
    pv["1D Ratio"] = (pv["1 Day"] / pv["Komplit"]) 
    pv["1W Ratio"] = (pv["1 Week"] / pv["Komplit"]) 
    pv["Cashless Ratio"] = (pv["Cashless"] / ( pv["Cashless"] + pv["Cash"]))
    pv["STT 30 VS OTS"] = (pv["STT 30"] / pv["OTS"]) 
    return pv

def calc_productifitas(dfData: pd.DataFrame, byMWC : bool = True):
    idx = ["Regional","Cabang", "Work Center"] if byMWC == True else ["Regional","Cabang", "Mn.wk.ctr", "Teknisi"]
    cfg = pd.read_excel(FILE_CONFIG, "seting").set_index("Seting")["Value"].to_dict()
    
    # pecah lagi data OTS & data Complete
    dfComplete = dfData[dfData["isOTS"].isin([0])]
    dfCash = dfData[dfData["UserStatus"].isin([93])]
    # Abaikan type ZX untuk perhitungan Cashless
    dfCash = dfCash[~dfCash["Typ"].isin(["ZX"])]

    # Pivoting
    pv = pd.pivot_table(dfComplete, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn": "Total LK"})
    pv_1D = pd.pivot_table(dfComplete, index=idx, values=["1D"], aggfunc="sum", fill_value=0, margins=True).rename(columns={"1D": "1 Day"})
    pv_1W = pd.pivot_table(dfComplete, index=idx, values=["1W"], aggfunc="sum", fill_value=0, margins=True).rename(columns={"1W": "1 Week"})
    pv_cash = pd.pivot_table(dfCash, index=idx, values=["Notifctn"], aggfunc="count", fill_value=0, margins=True).rename(columns={"Notifctn":"Cash"})
    pv_cashless = pd.pivot_table(dfComplete, index=idx, values=["Cashless"], aggfunc="sum", fill_value=0, margins=True)
    pv_TAT = pd.pivot_table(dfComplete, index=idx, values=["Lo"], aggfunc="mean", fill_value=0, margins=True).rename(columns={"Lo": "TAT"})
    pv = reduce(lambda left, right: pd.merge(left, right, on=idx, how="left"), [pv, pv_1D, pv_1W, pv_cash, pv_cashless, pv_TAT]).fillna(0)

    pv["Produktifitas"] = pv["Total LK"] / cfg["Hari"]
    pv["1D Ratio"] = (pv["1 Day"] / pv["Total LK"]) 
    pv["1W Ratio"] = (pv["1 Week"] / pv["Total LK"]) 
    pv["Cashless Ratio"] = (pv["Cashless"] / ( pv["Cashless"] + pv["Cash"]))
    return pv

def apply_filter(dfData: pd.DataFrame, CDR = "Cabang"):
    cfg = pd.read_excel(FILE_CONFIG, "seting").set_index("Seting")["Value"].to_dict()
    for col in ["Regional", "Cabang"]:
        if cfg.get(col, "All") != "All":
            dfData = dfData[dfData[col] == cfg[col]]

    dfData = dfData[dfData["CDR"] == CDR]
    return dfData

def format_result(table_pos: dict):
    try:
        wb = openpyxl.load_workbook(FILE_RESULT)
    except Exception as e:
        msg = f"Gagal membuka file {FILE_RESULT}: {e}"
        print(msg)
        raise SystemExit(msg)

    cfg = pd.read_excel(FILE_CONFIG, "bp")
    config = {
        row["item"]: {"mode": row["kondisi"], "value": row["bp"]}
        for _, row in cfg.iterrows()
    }

    #print(table_pos)

    # Gaya umum
    border = Border(
        bottom=Side(style='thin'), top=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_ng = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sky_blue = PatternFill(fill_type='solid', fgColor='FFC0FFC0')

    #print(excel_tables)
    
    col_map = {
        'nasional': {'LO': 9, 'TAT': 10, 'Cplt Ratio': 11, 'LO Ratio': 12, '1D Ratio': 13, '1W Ratio': 14, 'Cashless Ratio': 15, 'STT 30 VS OTS': 16},
        'Cabang': {'LO': 10, 'TAT': 11, 'Cplt Ratio': 12, 'LO Ratio': 13, '1D Ratio': 14, '1W Ratio': 15, 'Cashless Ratio': 16, 'STT 30 VS OTS': 17},
        'SDSS': {'LO': 10, 'TAT': 11, 'Cplt Ratio': 12, 'LO Ratio': 13, '1D Ratio': 14, '1W Ratio': 15, 'Cashless Ratio': 16, 'STT 30 VS OTS': 17},
        'SSR': {'LO': 10, 'TAT': 11, 'Cplt Ratio': 12, 'LO Ratio': 13, '1D Ratio': 14, '1W Ratio': 15, 'Cashless Ratio': 16, 'STT 30 VS OTS': 17},
        'ByMWC': {'TAT': 8, 'Produktifitas': 9 , '1D Ratio': 10, '1W Ratio': 11, 'Cashless Ratio': 12}, 
        'ByTech': {'TAT': 9, 'Produktifitas': 10 , '1D Ratio': 11, '1W Ratio': 12, 'Cashless Ratio': 13},
        'ssr': {'TAT': 8, 'Produktifitas': 9 , '1D Ratio': 10, '1W Ratio': 11, 'Cashless Ratio': 12},
    }

    for sheet in wb.worksheets:
        # Loop untuk setiap tabel dalam setiap sheet
        for table, coord in table_pos[sheet.title].items():
            # Format Header
            for col in sheet.iter_cols(
                min_row=coord["start_row"], max_row=coord["start_row"],
                min_col=coord["start_col"], max_col=coord["end_col"]
            ):
                 for idx, cell in enumerate(col):
                     cell.fill = sky_blue
                     cell.alignment = align_center
                     cell.border = border

            # Format Footer
            for col in sheet.iter_cols(
                min_row=coord["end_row"], max_row=coord["end_row"],
                min_col=coord["start_col"], max_col=coord["end_col"]
            ):
                 for idx, cell in enumerate(col):
                     cell.fill = sky_blue
                     cell.alignment = align_center
                     cell.border = border

            # Format body tabel
            for row in sheet.iter_rows(
                min_row=coord["start_row"]+1, max_row=coord["end_row"],
                min_col=coord["start_col"], max_col=coord["end_col"]
            ):
                for idx, cell in enumerate(row):
                    for i, val in col_map[table].items():
                        if table == 'nasional' :
                            judul = 1
                        elif table == 'ByTech':
                            judul = 4
                        else:
                            judul = 3

                        if idx <= judul:    
                            cell.alignment = align_left
                        
                        if idx == col_map[table][i]:
                            cell.alignment = align_center
                            cell.border = border
                            cell.number_format = '0.00' if i in ["LO", "TAT", "Produktifitas"] else '0.00%'                            
                            cell_value = float(cell.value)
                            if i in config:
                                if config[i]["mode"] == 'min':
                                    cell.fill = fill_ok if cell_value >= config[i]["value"] else fill_ng
                                else: 
                                    cell.fill = fill_ok if cell_value <= config[i]["value"] else fill_ng
                            else:
                                continue
                        else:
                            cell.border = border
    wb.save(FILE_RESULT)

# --- Proses Utama ---

# Cek file konfigurasi wajib
#check_file_exists(FILE_CONFIG)

# Cek file sumber lainnya 
#check_file_exists(FILE_OTS_ALT)
#check_file_exists(FILE_COMPLETED_ALT)

# Hapus file hasil jika sudah ada
for file in [FILE_RESULT, FILE_DETAIL, FILE_ERROR]:
    safe_remove(file)

# Load Source file & merge
df_ots = load_source(check_file(FILE_OTS_ALT, True))
df_completed = load_source(check_file(FILE_COMPLETED_ALT, True))
source = pd.concat([df_ots, df_completed])

# Hitung pencapaian global secara general
nasional = calc_achivement(source, True)


# Hitung pencapaian per kategori
result = {
    "Cabang": calc_achivement(apply_filter(source)).fillna(0),
    "SDSS": calc_achivement(apply_filter(source, "SDSS"), False, False).fillna(0),
    "SSR": calc_achivement(apply_filter(source, "SSR"), False, False).fillna(0)
}

result_tech = {
    "ByMWC": calc_productifitas(apply_filter(source)).fillna(0),
    "ByTech": calc_productifitas(apply_filter(source), False).fillna(0)
}

result_SDSS = {
    "ByMWC": calc_productifitas(apply_filter(source, "SDSS")).fillna(0),
    "ByTech": calc_productifitas(apply_filter(source, "SDSS"), False).fillna(0)
}

result_SSR = calc_productifitas(apply_filter(source, "SSR")).fillna(0)
excel_tables = {}
# Simpan ke Excel
with pd.ExcelWriter(FILE_RESULT, engine="openpyxl") as writer:
    row = 0
    nasional.to_excel(writer, sheet_name=SHEET_NAMES["Result"], startrow=row)
    row += len(nasional) + 5
    excel_tables[SHEET_NAMES["Result"]] = {
        "nasional" : {
            "start_row": 1,
            "end_row": len(nasional)+1,
            "start_col": 1,
            "end_col": len(nasional.columns)+1
        }
    }

    for label, df in result.items():
        df.to_excel(writer, sheet_name=SHEET_NAMES["Result"], startrow=row)
        start_row = row
        row += len(df) + 5
        excel_tables[SHEET_NAMES["Result"]].update({
            label: {
                "start_row": start_row+1,
                "end_row" : len(df)+1 + start_row,
                "start_col": 1,
                "end_col": len(df.columns)+2
            }
        })

    col = 0
    excel_tables[SHEET_NAMES["Prod_C"]] = {}
    for label, df in result_tech.items():
        shift = 3 if label == 'ByMWC' else 4
        excel_tables[SHEET_NAMES["Prod_C"]].update({
            label: {
                "start_row": 1,
                "end_row": len(df)+1,
                "start_col": col+1,
                "end_col": len(df.columns)+col+shift,
            }
        })
        df.to_excel(writer, sheet_name=SHEET_NAMES["Prod_C"], startcol=col, startrow=0)
        col += len(df.columns) + 5
        

    col = 0
    excel_tables[SHEET_NAMES["Prod_D"]] = {}
    for label, df in result_SDSS.items():
        shift = 3 if label == 'ByMWC' else 4
        excel_tables[SHEET_NAMES["Prod_D"]].update({
            label: {
                "start_row": 1,
                "end_row": len(df)+1,
                "start_col": col+1,
                "end_col": len(df.columns)+col+shift
            }
        })
        df.to_excel(writer, sheet_name=SHEET_NAMES["Prod_D"], startcol=col, startrow=0)
        col += len(df.columns) + 5
        

    col = 0
    excel_tables[SHEET_NAMES["Prod_R"]] = {}
    result_SSR.to_excel(writer, sheet_name=SHEET_NAMES["Prod_R"], startcol=col, startrow=0)
    excel_tables[SHEET_NAMES["Prod_R"]].update({
            "ssr": {
                "start_row": 1,
                "end_row": len(result_SSR)+1,
                "start_col": col+1,
                "end_col": len(result_SSR.columns)+3,
            }
        })

# Format Result
format_result(excel_tables)