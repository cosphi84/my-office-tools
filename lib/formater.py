import openpyxl
from openpyxl.styles import Border, Side, Alignment, PatternFill
from pathlib import Path
import pandas as pd
from Config.config import csopp_config

def format_result(table_pos: dict):
    file: dict = csopp_config().get('csopp_files')
    path = file.get('FILE_RESULT')

    
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        msg = f"Gagal membuka file {path}: {e}"
        print(msg)
        raise SystemExit(msg)

    
    cfg = csopp_config().get('csopp_bp')
    config = {
        row["item"]: {"mode": row["kondisi"], "value": row["bp"]}
        for _, row in cfg.iterrows()
    }

    #print(table_pos)

    # Gaya umum
    border = Border(
        bottom=Side(style='thin'), top=Side(style='thin'),
        left=Side(style='thin'), right=Side(style='thin')
    )
    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')
    fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_ng = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sky_blue = PatternFill(fill_type='solid', fgColor='0099CCFF')

    #print(excel_tables)
    col_maps = {
        "Pencapaian" : {
            "TAT": 6,
            "Cplt Ratio": 11,
            "1D Ratio": 12,
            "1W Ratio": 13,
            "Cashless Ratio": 14,
            "LO Ratio": 15,
            "STT 30 VS OTS": 16,
            "STT 30 VS ALL": 17
        },
        "Produktifitas": {
            "TAT": 2,
            "Produktifitas": 7,
            "1D Ratio": 8,
            "1W Ratio": 9,
            "Cashless Ratio": 10
        }
    }

    for sheet in wb.worksheets:
        # Loop untuk setiap tabel dalam setiap sheet
        
        for table, coord in table_pos[sheet.title].items():
            # Panjang Index Table as constant
            table_index = coord['nindex']-1
            # Format Header
            for col in sheet.iter_cols(
                min_row=coord["start_row"]+1, max_row=coord["start_row"]+1,
                min_col=coord["start_col"]+1, max_col=coord["end_col"]+2+table_index
            ):
                 for idx, cell in enumerate(col):
                     cell.fill = sky_blue
                     cell.alignment = align_center
                     cell.border = border

            # Format body tabel
            for row in sheet.iter_rows(
                min_row=coord["start_row"]+1, max_row=coord["end_row"]+1,
                min_col=coord["start_col"]+1, max_col=coord["end_col"]+2+table_index
            ):
                for idx, cell in enumerate(row):
                    cell.border = border
                    
                    locs = col_maps["Pencapaian"] if sheet.title == "Pencapaian" else col_maps["Produktifitas"]

                    for i,loc in locs.items():
                        if idx == (loc + table_index):
                            cell.alignment = align_center
                            cell.number_format = '0.00' if i in ["TAT", "Produktifitas"] else '0.00%'
                            try:
                                val = float(cell.value)                                           
                                if i in config:
                                    if config[i]["mode"] == 'min':
                                        cell.fill = fill_ok if val >= config[i]["value"] else fill_ng
                                    else: 
                                        cell.fill = fill_ok if val <= config[i]["value"] else fill_ng
                                else:
                                    continue
                            except (TypeError, ValueError):                    
                                continue
                        elif idx <= table_index:
                            cell.alignment = align_left
                        else:
                            cell.alignment = align_center
    wb.save(path)